from sqlalchemy import or_
from datetime import datetime
from flask import current_app
from app.models import db, Document, Contact, DocumentType


def build_invoice_query(company_id, filters):
    search = filters.get("search", "")
    status = filters.get("status", "")
    doc_type = filters.get("type", "")
    date_from = filters.get("date_from", "")
    date_to = filters.get("date_to", "")

    query = db.session.query(Document).filter(
        Document.company_id == company_id,
        or_(Document.type == DocumentType.invoice,
            Document.type == DocumentType.quote)
    )

    if search:
        query = query.outerjoin(Contact).filter(
            or_(
                Document.document_number.ilike(f"%{search}%"),
                Contact.name.ilike(f"%{search}%")
            )
        )

    if status:
        query = query.filter(Document.status == status)

    if doc_type == "invoice":
        query = query.filter(Document.type == DocumentType.invoice)
    elif doc_type == "quote":
        query = query.filter(Document.type == DocumentType.quote)

    if date_from:
        try:
            query = query.filter(
                Document.issued_date >= datetime.strptime(date_from, "%Y-%m-%d")
            )
        except ValueError:
            current_app.logger.warning(f"Invalid date_from format: {date_from}")

    if date_to:
        try:
            query = query.filter(
                Document.issued_date <= datetime.strptime(date_to, "%Y-%m-%d")
            )
        except ValueError:
            current_app.logger.warning(f"Invalid date_to format: {date_to}")

    query = query.order_by(Document.id.desc())
    return query


def get_invoice_list(company_id, filters):
    page = int(filters.get("page", 1))
    query = build_invoice_query(company_id, filters)

    pagination = query.paginate(
        page=page,
        per_page=int(current_app.config.get("ITEMS_PER_PAGE", 20)),
        error_out=False
    )

    return pagination


def export_invoice_report_xlsx(company_id, filters):
    """Return (workbook, filename) for all invoice rows matching the active filters."""
    from datetime import UTC
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    query = build_invoice_query(company_id, filters)
    documents = query.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Facturas"

    headers = [
        "N° Documento",
        "Cliente",
        "Tipo",
        "Estado",
        "Fecha Emisión",
        "Fecha Vencimiento",
        "Subtotal",
        "Impuesto",
        "Total",
        "Pagado",
        "Saldo",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="E2F5EC")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="14532D")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    type_labels = {
        "invoice": "Factura",
        "quote": "Cotización",
    }
    status_labels = {
        "draft": "Borrador",
        "sent": "Enviada",
        "issued": "Emitida",
        "partial": "Parcial",
        "paid": "Pagada",
        "pending": "Pendiente",
        "overdue": "Vencida",
        "cancelled": "Cancelada",
        "credit_note": "Nota de Crédito",
        "exchange": "Intercambio",
    }

    for document in documents:
        subtotal = float(document.subtotal_cache or document.subtotal or 0)
        tax = float(document.tax_cache or document.tax_amount or 0)
        total = float(document.total_amount or 0)
        paid = document.calculate_paid_amount()
        balance = document.calculate_balance_due()
        doc_type = document.type.value if document.type else ""
        status = document.status.value if document.status else ""

        ws.append([
            document.document_number,
            document.client.name if document.client else "",
            type_labels.get(doc_type, doc_type),
            status_labels.get(status, status),
            document.issued_date.strftime("%Y-%m-%d") if document.issued_date else "",
            document.due_date.strftime("%Y-%m-%d") if document.due_date else "",
            subtotal,
            tax,
            total,
            paid,
            balance,
        ])

    for row in ws.iter_rows(min_row=2, min_col=7, max_col=11):
        for cell in row:
            cell.number_format = '#,##0.00'

    widths = [18, 28, 14, 16, 16, 18, 14, 14, 14, 14, 14]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    status = filters.get("status") or "todos"
    doc_type = filters.get("type") or "documentos"
    date_stamp = datetime.now(UTC).strftime("%Y%m%d")
    filename = f"facturas_{doc_type}_{status}_{date_stamp}.xlsx"
    return wb, filename
