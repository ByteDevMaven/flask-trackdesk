from datetime import datetime, timedelta, UTC

from sqlalchemy import or_, and_, desc, asc, func
from flask_babel import _
from flask_login import current_user

from app.models import db, InventoryItem, StockMovement, StockMovementType


class InventoryService:
    @staticmethod
    def get_inventory_items(company_id, page=1, per_page=15, search='', supplier_id=None, sort_by='name', sort_order='asc'):
        query = InventoryItem.query.filter_by(company_id=company_id)
        
        if search:
            query = query.filter(
                or_(
                    InventoryItem.name.ilike(f'%{search}%'),
                    InventoryItem.description.ilike(f'%{search}%')
                )
            )
        
        if supplier_id and str(supplier_id).isdigit():
            query = query.filter_by(supplier_id=int(supplier_id))
            
        if sort_by == 'name':
            query = query.order_by(asc(InventoryItem.name) if sort_order == 'asc' else desc(InventoryItem.name))
        elif sort_by == 'quantity':
            query = query.order_by(asc(InventoryItem.quantity) if sort_order == 'asc' else desc(InventoryItem.quantity))
        elif sort_by == 'price':
            query = query.order_by(asc(InventoryItem.price) if sort_order == 'asc' else desc(InventoryItem.price))
        else:
            query = query.order_by(InventoryItem.name)
            
        return query.paginate(page=page, per_page=per_page, error_out=False)

    @staticmethod
    def get_inventory_stats(company_id):
        total_items = InventoryItem.query.filter_by(company_id=company_id).count()
        low_stock_items = InventoryItem.query.filter(
            and_(InventoryItem.company_id == company_id, InventoryItem.quantity <= 10)
        ).count()
        out_of_stock_items = InventoryItem.query.filter(
            and_(InventoryItem.company_id == company_id, InventoryItem.quantity == 0)
        ).count()
        total_value = db.session.query(func.sum(InventoryItem.quantity * InventoryItem.price))\
            .filter_by(company_id=company_id).scalar() or 0
        
        return {
            'total_items': total_items,
            'low_stock_items': low_stock_items,
            'out_of_stock_items': out_of_stock_items,
            'total_value': float(total_value),
            'in_stock_items': total_items - out_of_stock_items
        }

    @staticmethod
    def adjust_stock(company_id, item_id, warehouse_id, adjustment, reference=None):
        item = InventoryItem.query.filter_by(id=item_id, company_id=company_id).first_or_404()
        from app.models import WarehouseItem
        warehouse_item = WarehouseItem.query.filter_by(warehouse_id=warehouse_id, inventory_item_id=item_id).first()
        if not warehouse_item:
            warehouse_item = WarehouseItem(warehouse_id=warehouse_id, inventory_item_id=item_id, quantity=0)
            db.session.add(warehouse_item)
            
        new_quantity = max(0, warehouse_item.quantity + adjustment)
        
        movement = StockMovement(
            company_id=company_id,
            inventory_item_id=item_id,
            warehouse_id=warehouse_id,
            user_id=current_user.id if current_user.is_authenticated else None,
            type=StockMovementType.adjustment,
            quantity=adjustment,
            reference=reference or _('Manual Adjustment'),
            date=datetime.now(UTC)
        )
        db.session.add(movement)
        
        warehouse_item.quantity = new_quantity
        item.quantity = max(0, item.quantity + adjustment)
        
        db.session.commit()
        return new_quantity

    @staticmethod
    def transfer_stock(company_id, item_id, from_warehouse_id, to_warehouse_id, quantity, reference=None):
        if quantity <= 0:
            raise ValueError(_("Quantity must be positive"))
            
        item = InventoryItem.query.filter_by(id=item_id, company_id=company_id).first_or_404()
        from app.models import WarehouseItem
        
        from_item = WarehouseItem.query.filter_by(warehouse_id=from_warehouse_id, inventory_item_id=item_id).first()
        if not from_item or from_item.quantity < quantity:
            raise ValueError(_("Not enough stock in source warehouse"))
            
        to_item = WarehouseItem.query.filter_by(warehouse_id=to_warehouse_id, inventory_item_id=item_id).first()
        if not to_item:
            to_item = WarehouseItem(warehouse_id=to_warehouse_id, inventory_item_id=item_id, quantity=0)
            db.session.add(to_item)
            
        from_item.quantity -= quantity
        to_item.quantity += quantity
        
        movement = StockMovement(
            company_id=company_id,
            inventory_item_id=item_id,
            warehouse_id=from_warehouse_id,
            destination_warehouse_id=to_warehouse_id,
            user_id=current_user.id if current_user.is_authenticated else None,
            type=StockMovementType.outgoing,
            quantity=quantity,
            reference=reference or _('Stock Transfer'),
            date=datetime.now(UTC)
        )
        db.session.add(movement)
        db.session.commit()
        return True


    @staticmethod
    def create_inventory_item(company_id, name, description=None, quantity=0, price=0.0, cost_price=0.0, discount=0.0, supplier_id=None, warehouse_id=None, sku=None):
        if not name:
            raise ValueError(_('Name is required'))
        if quantity < 0:
            raise ValueError(_('Quantity cannot be negative'))
        if price < 0:
            raise ValueError(_('Price cannot be negative'))
        if cost_price < 0:
            raise ValueError(_('Cost price cannot be negative'))
        if discount < 0 or discount > 100:
            raise ValueError(_('Discount must be between 0 and 100'))

        item = InventoryItem(
            company_id=company_id,
            name=name,
            description=description,
            quantity=quantity,
            price=price,
            cost_price=cost_price,
            discount=discount,
            supplier_id=int(supplier_id) if supplier_id and str(supplier_id).isdigit() else None
        )
        db.session.add(item)
        db.session.flush() # flush to get item.id

        # Generate or validate SKU
        if sku and sku.strip():
            candidate = sku.strip().upper()
            existing = InventoryItem.query.filter_by(company_id=company_id, sku=candidate).first()
            if existing and existing.id != item.id:
                raise ValueError(_('A product with this SKU already exists'))
            item.sku = candidate
        else:
            item.sku = InventoryItem.build_sku(name, item.id)
        
        if quantity > 0:
            if not warehouse_id:
                from app.models import Warehouse
                default_warehouse = Warehouse.query.filter_by(company_id=company_id).first()
                if default_warehouse:
                    warehouse_id = default_warehouse.id
            if warehouse_id:
                from app.models import WarehouseItem
                wh_item = WarehouseItem(warehouse_id=warehouse_id, inventory_item_id=item.id, quantity=quantity)
                db.session.add(wh_item)
                
                movement = StockMovement(
                    company_id=company_id,
                    inventory_item_id=item.id,
                    warehouse_id=warehouse_id,
                    user_id=current_user.id if current_user.is_authenticated else None,
                    type=StockMovementType.incoming,
                    quantity=quantity,
                    reference=_('Initial Stock'),
                    date=datetime.now(UTC)
                )
                db.session.add(movement)

        db.session.commit()
        return item

    @staticmethod
    def get_inventory_item(company_id, item_id):
        return InventoryItem.query.filter_by(id=item_id, company_id=company_id).first()

    @staticmethod
    def get_item_by_sku(company_id, sku):
        """Fetch an item by its SKU within a company."""
        return InventoryItem.query.filter_by(company_id=company_id, sku=sku).first()

    @staticmethod
    def update_inventory_item(company_id, item_id, name=None, description=None, quantity=None, price=None, cost_price=None, discount=None, supplier_id=None, sku=None):
        item = InventoryItem.query.filter_by(id=item_id, company_id=company_id).first_or_404()
        
        if name is not None:
            if not name.strip():
                raise ValueError(_('Name is required'))
            item.name = name.strip()
            
        if description is not None:
            item.description = description if description else None
            
        if quantity is not None:
            if quantity < 0:
                raise ValueError(_('Quantity cannot be negative'))
            item.quantity = quantity
            
        if price is not None:
            if price < 0:
                raise ValueError(_('Price cannot be negative'))
            item.price = price

        if cost_price is not None:
            if cost_price < 0:
                raise ValueError(_('Cost price cannot be negative'))
            item.cost_price = cost_price

        if discount is not None:
            if discount < 0 or discount > 100:
                raise ValueError(_('Discount must be between 0 and 100'))
            item.discount = discount
            
        if supplier_id is not None:
            item.supplier_id = int(supplier_id) if str(supplier_id).isdigit() else None

        if sku is not None:
            candidate = sku.strip().upper() if sku.strip() else InventoryItem.build_sku(item.name, item.id)
            existing = InventoryItem.query.filter_by(company_id=company_id, sku=candidate).first()
            if existing and existing.id != item.id:
                raise ValueError(_('A product with this SKU already exists'))
            item.sku = candidate
            
        db.session.commit()
        return item

    @staticmethod
    def delete_inventory_item(company_id, item_id):
        item = InventoryItem.query.filter_by(id=item_id, company_id=company_id).first_or_404()
        item.is_deleted = True
        item.deleted_at = datetime.now(UTC)
        db.session.commit()
        return True

    @staticmethod
    def get_stock_movements(company_id, item_id=None, movement_type=None, period='all', search='', client_id=None, supplier_id=None, page=1, per_page=20):
        query = StockMovement.query.filter_by(company_id=company_id)
        
        if item_id:
            query = query.filter_by(inventory_item_id=item_id)
            
        if client_id or supplier_id:
            from app.models import PurchaseOrder, Document
            # Use outer join to allow filtering by contact_id based on reference
            query = query.outerjoin(
                PurchaseOrder,
                StockMovement.reference == ("PO " + PurchaseOrder.order_number)
            ).outerjoin(
                Document,
                StockMovement.reference == ("INV " + Document.document_number)
            )
            
            conditions = []
            if supplier_id:
                conditions.append(PurchaseOrder.supplier_id == supplier_id)
            if client_id:
                conditions.append(Document.client_id == client_id)
            
            query = query.filter(or_(*conditions) if len(conditions) > 1 else conditions[0])

        if search:
            query = query.join(InventoryItem).filter(
                or_(
                    InventoryItem.name.ilike(f'%{search}%'),
                    StockMovement.reference.ilike(f'%{search}%')
                )
            )
            
        if movement_type and movement_type in [t.value for t in StockMovementType]:
            query = query.filter(StockMovement.type == movement_type)
            
        today = datetime.now()
        if period == 'day':
            start_date = today.replace(hour=0, minute=0, second=0, microsecond=0)
            query = query.filter(StockMovement.date >= start_date)
        elif period == 'week':
            start_date = (today - timedelta(days=today.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
            query = query.filter(StockMovement.date >= start_date)
        elif period == 'month':
            start_date = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            query = query.filter(StockMovement.date >= start_date)
            
        return query.order_by(StockMovement.date.desc()).paginate(page=page, per_page=per_page, error_out=False)

    @staticmethod
    def bulk_delete_items(company_id, item_ids):
        if not item_ids:
            return 0
        
        items = InventoryItem.query.filter(
            and_(InventoryItem.id.in_(item_ids), InventoryItem.company_id == company_id)
        ).all()
        for item in items:
            item.is_deleted = True
            item.deleted_at = datetime.now(UTC)
        
        db.session.commit()
        return len(items)

    @staticmethod
    def search_inventory_items(company_id, query, limit=10):
        if not query or len(query) < 2:
            return []
            
        return InventoryItem.query.filter(
            and_(
                InventoryItem.company_id == company_id,
                or_(
                    InventoryItem.name.ilike(f'%{query}%'),
                    InventoryItem.description.ilike(f'%{query}%')
                )
            )
        ).limit(limit).all()

    @staticmethod
    def export_inventory_items_xlsx(company_id, search='', supplier_id=None):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from flask_babel import _
        
        pagination = InventoryService.get_inventory_items(
            company_id=company_id,
            page=1,
            per_page=1000000,
            search=search,
            supplier_id=supplier_id
        )
        items = pagination.items
        
        wb = Workbook()
        ws = wb.active
        ws.title = _('Inventario')
        
        headers = [_('Nombre'), _('Descripción'), _('Cantidad'), _('Precio'), _('Costo'), _('Proveedor')]
        ws.append(headers)
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        for item in items:
            ws.append([
                item.name,
                item.description or '',
                item.quantity,
                float(item.price),
                float(item.cost_price),
                item.supplier.name if item.supplier else ''
            ])
            
        filename = f"inventario_{company_id}_{datetime.now(UTC).strftime('%Y%m%d')}.xlsx"
        return wb, filename

    @staticmethod
    def export_stock_movements_xlsx(company_id, movement_type=None, period='all', search='', client_id=None, supplier_id=None):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from flask_babel import _
        
        # Get all movements without pagination
        pagination = InventoryService.get_stock_movements(
            company_id=company_id,
            movement_type=movement_type,
            period=period,
            search=search,
            client_id=client_id,
            supplier_id=supplier_id,
            page=1,
            per_page=1000000
        )
        movements = pagination.items
        
        wb = Workbook()
        ws = wb.active
        ws.title = _('Movimientos de Inventario')
        
        headers = [_('Fecha'), _('Tipo'), _('Producto'), _('Referencia'), _('Cantidad'), _('Notas')]
        ws.append(headers)
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        for m in movements:
            type_str = ''
            if m.type == StockMovementType.incoming:
                type_str = _('Entrada')
            elif m.type == StockMovementType.outgoing:
                type_str = _('Salida')
            else:
                type_str = _('Ajuste')
                
            date_str = m.date.strftime('%Y-%m-%d %H:%M') if m.date else ''
            prod_name = m.inventory_item.name if m.inventory_item else ''
            
            ws.append([
                date_str,
                type_str,
                prod_name,
                m.reference or '',
                m.qty_change,
                m.notes or ''
            ])
            
        filename = f"movimientos_inventario_{company_id}_{datetime.now(UTC).strftime('%Y%m%d')}.xlsx"
        return wb, filename
